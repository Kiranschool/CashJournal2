from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView, View
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Sum, Count
from django.utils import timezone
from datetime import datetime, timedelta
from django.contrib.auth import login, logout
from django.http import JsonResponse, HttpResponse
from .models import Transaction, JournalEntry, WishlistItem, Category, UserPreferences
from .forms import TransactionForm, JournalEntryForm, WishlistItemForm, UserSignupForm
from .utils import convert_currency, format_currency
from .nlp_utils import get_category_suggestions, get_similar_transactions
from .forecast_utils import predict_future_spending
from django.views.decorators.csrf import ensure_csrf_cookie
import json

# Try to import openpyxl, but don't fail if it's not available
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class HomeView(LoginRequiredMixin, ListView):
    template_name = 'core/home.html'
    context_object_name = 'transactions'

    def get_queryset(self):
        # Get recent transactions
        return Transaction.objects.filter(user=self.request.user).order_by('-date')[:5]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Ensure user preferences exist
        user_preferences, created = UserPreferences.objects.get_or_create(user=self.request.user)
        user_currency = user_preferences.default_currency
        
        # Get monthly summary
        today = timezone.now()
        first_day = today.replace(day=1)
        last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        monthly_expenses = Transaction.objects.filter(
            user=self.request.user,
            date__range=[first_day, last_day],
            transaction_type='expense'
        ).aggregate(total=Sum('amount'))['total'] or 0

        monthly_income = Transaction.objects.filter(
            user=self.request.user,
            date__range=[first_day, last_day],
            transaction_type='income'
        ).aggregate(total=Sum('amount'))['total'] or 0

        # Convert amounts to user's preferred currency
        monthly_expenses = convert_currency(monthly_expenses, 'USD', user_currency)
        monthly_income = convert_currency(monthly_income, 'USD', user_currency)

        context['monthly_summary'] = {
            'expenses': monthly_expenses,
            'income': monthly_income,
            'balance': monthly_income - monthly_expenses,
            'currency': user_currency
        }
        context['month'] = today.strftime('%B %Y')

        # Get category breakdown
        category_spending = Transaction.objects.filter(
            user=self.request.user,
            date__range=[first_day, last_day],
            transaction_type='expense'
        ).values('category__name').annotate(total=Sum('amount'))

        # Convert category totals to user's preferred currency
        for category in category_spending:
            category['total'] = convert_currency(category['total'], 'USD', user_currency)

        context['category_breakdown'] = category_spending
        return context

class FinanceCalendarView(LoginRequiredMixin, ListView):
    template_name = 'core/finance_calendar.html'
    context_object_name = 'transactions'

    def get_queryset(self):
        year = int(self.request.GET.get('year', timezone.now().year))
        month = int(self.request.GET.get('month', timezone.now().month))
        return Transaction.objects.filter(
            user=self.request.user,
            date__year=year,
            date__month=month
        ).order_by('date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Ensure user preferences exist
        user_preferences, created = UserPreferences.objects.get_or_create(user=self.request.user)
        user_currency = user_preferences.default_currency
        
        year = int(self.request.GET.get('year', timezone.now().year))
        month = int(self.request.GET.get('month', timezone.now().month))
        
        # Handle month transitions
        if month > 12:
            year += 1
            month = 1
        elif month < 1:
            year -= 1
            month = 12
        
        # Get the first day of the month
        first_day = datetime(year, month, 1)
        # Get the last day of the month
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)
        
        # Create calendar data
        calendar = []
        week = []
        
        # Add empty days for the first week
        first_weekday = first_day.weekday()
        for _ in range(first_weekday):
            week.append({'day': '', 'is_current_month': False, 'transactions': []})
        
        # Add days of the month
        current_day = first_day
        while current_day <= last_day:
            if len(week) == 7:
                calendar.append(week)
                week = []
            
            # Get transactions for this day
            day_transactions = self.get_queryset().filter(date=current_day)
            
            # Convert transaction amounts to user's preferred currency
            for transaction in day_transactions:
                transaction.converted_amount = convert_currency(transaction.amount, 'USD', user_currency)
            
            week.append({
                'day': current_day.day,
                'is_current_month': True,
                'is_today': current_day.date() == timezone.now().date(),
                'transactions': day_transactions
            })
            
            current_day += timedelta(days=1)
        
        # Add empty days for the last week
        while len(week) < 7:
            week.append({'day': '', 'is_current_month': False, 'transactions': []})
        calendar.append(week)
        
        context['calendar'] = calendar
        context['year'] = year
        context['month'] = month
        context['month_name'] = datetime(year, month, 1).strftime('%B')
        context['current_year'] = timezone.now().year
        context['current_month'] = timezone.now().month
        context['currency'] = user_currency
        return context

class TransactionCreateView(LoginRequiredMixin, CreateView):
    model = Transaction
    form_class = TransactionForm
    template_name = 'core/transaction_form.html'
    success_url = reverse_lazy('core:finance_calendar')

    def get_initial(self):
        initial = super().get_initial()
        date_str = self.request.GET.get('date')
        if date_str:
            try:
                initial['date'] = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        return initial

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_currency'] = self.request.user.userpreferences.default_currency
        return context

    def form_valid(self, form):
        form.instance.user = self.request.user
        response = super().form_valid(form)
        return response

class TransactionUpdateView(LoginRequiredMixin, UpdateView):
    model = Transaction
    form_class = TransactionForm
    template_name = 'core/transaction_form.html'
    success_url = reverse_lazy('core:finance_calendar')

    def get_queryset(self):
        return Transaction.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_currency'] = self.request.user.userpreferences.default_currency
        return context

    def get_initial(self):
        initial = super().get_initial()
        # Convert amount from USD to user's currency for display
        if self.object:
            user_currency = self.request.user.userpreferences.default_currency
            if user_currency != 'USD':
                initial['amount'] = convert_currency(self.object.amount, 'USD', user_currency)
        return initial

class TransactionDeleteView(LoginRequiredMixin, DeleteView):
    model = Transaction
    success_url = reverse_lazy('core:finance_calendar')
    template_name = 'core/transaction_confirm_delete.html'

    def get_queryset(self):
        return Transaction.objects.filter(user=self.request.user)

class WishlistView(LoginRequiredMixin, ListView):
    model = WishlistItem
    template_name = 'core/wishlist.html'
    context_object_name = 'wishlist_items'

    def get_queryset(self):
        queryset = WishlistItem.objects.filter(user=self.request.user).order_by('priority')
        
        # Ensure user preferences exist
        user_preferences, created = UserPreferences.objects.get_or_create(user=self.request.user)
        user_currency = user_preferences.default_currency
        
        # Convert estimated costs to user's preferred currency
        for item in queryset:
            if user_currency != 'USD':
                item.converted_cost = convert_currency(item.estimated_cost, 'USD', user_currency)
            else:
                item.converted_cost = item.estimated_cost
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Ensure user preferences exist
        user_preferences, created = UserPreferences.objects.get_or_create(user=self.request.user)
        context['currency'] = user_preferences.default_currency
        return context

class WishlistItemCreateView(LoginRequiredMixin, CreateView):
    model = WishlistItem
    form_class = WishlistItemForm
    template_name = 'core/wishlist_item_form.html'
    success_url = reverse_lazy('core:wishlist')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_currency'] = self.request.user.userpreferences.default_currency
        return context

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)

class WishlistItemUpdateView(LoginRequiredMixin, UpdateView):
    model = WishlistItem
    form_class = WishlistItemForm
    template_name = 'core/wishlist_item_form.html'
    success_url = reverse_lazy('core:wishlist')

    def get_queryset(self):
        return WishlistItem.objects.filter(user=self.request.user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_currency'] = self.request.user.userpreferences.default_currency
        return context

    def get_initial(self):
        initial = super().get_initial()
        # Convert amount from USD to user's currency for display
        if self.object:
            user_currency = self.request.user.userpreferences.default_currency
            if user_currency != 'USD':
                initial['estimated_cost'] = convert_currency(self.object.estimated_cost, 'USD', user_currency)
        return initial

class WishlistItemDeleteView(LoginRequiredMixin, DeleteView):
    model = WishlistItem
    success_url = reverse_lazy('core:wishlist')
    template_name = 'core/wishlist_item_confirm_delete.html'

    def get_queryset(self):
        return WishlistItem.objects.filter(user=self.request.user)

class SpendingTrackerView(LoginRequiredMixin, ListView):
    template_name = 'core/spending_tracker.html'
    context_object_name = 'transactions'

    def get_queryset(self):
        return Transaction.objects.filter(user=self.request.user).order_by('-date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_currency = self.request.user.userpreferences.default_currency
        
        # Get current month's date range
        today = timezone.now()
        first_day = today.replace(day=1)
        last_day = (first_day + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Get spending by category
        category_spending = Transaction.objects.filter(
            user=self.request.user,
            transaction_type='expense'
        ).values('category__name').annotate(total=Sum('amount'))

        # Convert category totals to user's preferred currency
        for category in category_spending:
            category['total'] = convert_currency(category['total'], 'USD', user_currency)

        context['category_spending'] = category_spending

        # Get monthly trend
        monthly_trend = Transaction.objects.filter(
            user=self.request.user,
            transaction_type='expense'
        ).values('date__year', 'date__month').annotate(total=Sum('amount')).order_by('date__year', 'date__month')

        # Convert monthly totals to user's preferred currency
        for month in monthly_trend:
            month['total'] = convert_currency(month['total'], 'USD', user_currency)

        context['monthly_trend'] = monthly_trend

        # Get purchase reflections for current month
        purchase_reflections = Transaction.objects.filter(
            user=self.request.user,
            date__range=[first_day, last_day],
            purchase_reflection__isnull=False
        ).values('purchase_reflection').annotate(count=Count('id'))

        # Format the data for the pie chart
        reflection_data = {
            'labels': [],
            'data': [],
            'colors': {
                'happy': '#4BC0C0',  # Teal
                'regret': '#FF6384',  # Red
                'undecided': '#FFCE56'  # Yellow
            }
        }

        for reflection in purchase_reflections:
            reflection_type = reflection['purchase_reflection']
            count = reflection['count']
            reflection_data['labels'].append(dict(Transaction.PURCHASE_REFLECTIONS)[reflection_type])
            reflection_data['data'].append(count)

        context['reflection_data'] = reflection_data

        # Get spending forecast
        all_transactions = Transaction.objects.filter(
            user=self.request.user,
            transaction_type='expense'
        ).order_by('date')
        
        forecast = predict_future_spending(all_transactions)
        
        # Convert forecast amounts to user's preferred currency
        for prediction in forecast:
            prediction['amount'] = convert_currency(prediction['amount'], 'USD', user_currency)
        
        context['spending_forecast'] = forecast
        context['currency'] = user_currency
        return context

class AccountSettingsView(LoginRequiredMixin, UpdateView):
    model = UserPreferences
    template_name = 'core/account_settings.html'
    fields = ['default_currency', 'notification_enabled']
    success_url = reverse_lazy('core:account_settings')

    def get_object(self):
        preferences, created = UserPreferences.objects.get_or_create(user=self.request.user)
        return preferences

    def form_valid(self, form):
        messages.success(self.request, 'Settings updated successfully!')
        return super().form_valid(form)

class SignupView(FormView):
    template_name = 'core/signup.html'
    form_class = UserSignupForm
    success_url = reverse_lazy('core:home')

    def form_valid(self, form):
        user = form.save()
        login(self.request, user)
        
        # Create default categories for the new user
        default_categories = [
            {'name': 'Food & Dining', 'color': '#FF5733'},
            {'name': 'Transportation', 'color': '#33FF57'},
            {'name': 'Shopping', 'color': '#3357FF'},
            {'name': 'Entertainment', 'color': '#FF33F5'},
            {'name': 'Bills & Utilities', 'color': '#33FFF5'},
            {'name': 'Health & Fitness', 'color': '#F5FF33'},
            {'name': 'Travel', 'color': '#FF8333'},
            {'name': 'Education', 'color': '#338FFF'},
            {'name': 'Personal Care', 'color': '#FF33A1'},
            {'name': 'Other', 'color': '#808080'}
        ]

        for category_data in default_categories:
            Category.objects.create(
                name=category_data['name'],
                color=category_data['color'],
                user=user,
                is_default=True
            )
        
        messages.success(self.request, 'Account created successfully!')
        return super().form_valid(form)

@ensure_csrf_cookie
def category_suggestions(request):
    """API endpoint for getting category suggestions"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    description = request.GET.get('description', '')
    if description:
        suggestions = get_category_suggestions(description)
        return JsonResponse({'suggestions': suggestions})
    return JsonResponse({'suggestions': []})

@ensure_csrf_cookie
def similar_transactions(request):
    """API endpoint for getting similar transactions"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    description = request.GET.get('description', '')
    if description:
        similar = get_similar_transactions(description)
        return JsonResponse({'similar_transactions': similar})
    return JsonResponse({'similar_transactions': []})

class ExportDataView(LoginRequiredMixin, View):
    def get(self, request):
        if not OPENPYXL_AVAILABLE:
            messages.error(request, "Excel export functionality is not available. Please contact support.")
            return redirect('core:home')
            
        # Create a new workbook and select the active sheet
        wb = Workbook()
        
        # Create sheets
        user_sheet = wb.active
        user_sheet.title = "User Info"
        transactions_sheet = wb.create_sheet("Transactions")
        wishlist_sheet = wb.create_sheet("Wishlist")
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # User Info Sheet
        user_sheet['A1'] = "User Information"
        user_sheet['A1'].font = header_font
        user_sheet['A1'].fill = header_fill
        user_sheet['A1'].border = border
        
        user_data = [
            ["Username", request.user.username],
            ["Email", request.user.email],
            ["Date Joined", request.user.date_joined.strftime('%Y-%m-%d')],
            ["Default Currency", request.user.userpreferences.default_currency],
            ["Notifications Enabled", "Yes" if request.user.userpreferences.notification_enabled else "No"]
        ]
        
        for row_idx, (label, value) in enumerate(user_data, start=2):
            user_sheet[f'A{row_idx}'] = label
            user_sheet[f'B{row_idx}'] = value
            user_sheet[f'A{row_idx}'].font = Font(bold=True)
            user_sheet[f'A{row_idx}'].border = border
            user_sheet[f'B{row_idx}'].border = border
        
        # Transactions Sheet
        transactions = Transaction.objects.filter(user=request.user).order_by('-date')
        transaction_headers = [
            "Date", "Amount", "Purchase Name", "Category", "Type",
            "Purchase Reflection", "Reflection", "Recurring", "Frequency", "End Date"
        ]
        
        for col_idx, header in enumerate(transaction_headers, start=1):
            cell = transactions_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, transaction in enumerate(transactions, start=2):
            transactions_sheet[f'A{row_idx}'] = transaction.date.strftime('%Y-%m-%d')
            transactions_sheet[f'B{row_idx}'] = float(transaction.amount)
            transactions_sheet[f'C{row_idx}'] = transaction.purchase_name
            transactions_sheet[f'D{row_idx}'] = transaction.category.name
            transactions_sheet[f'E{row_idx}'] = transaction.transaction_type
            transactions_sheet[f'F{row_idx}'] = transaction.purchase_reflection
            transactions_sheet[f'G{row_idx}'] = transaction.reflection
            transactions_sheet[f'H{row_idx}'] = "Yes" if transaction.is_recurring else "No"
            transactions_sheet[f'I{row_idx}'] = transaction.recurring_frequency if transaction.is_recurring else ""
            transactions_sheet[f'J{row_idx}'] = transaction.recurring_end_date.strftime('%Y-%m-%d') if transaction.is_recurring and transaction.recurring_end_date else ""
            
            for col in range(1, 11):
                transactions_sheet.cell(row=row_idx, column=col).border = border
        
        # Wishlist Sheet
        wishlist_items = WishlistItem.objects.filter(user=request.user).order_by('-created_at')
        wishlist_headers = ["Name", "Description", "Estimated Cost", "Priority", "Created At"]
        
        for col_idx, header in enumerate(wishlist_headers, start=1):
            cell = wishlist_sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, item in enumerate(wishlist_items, start=2):
            wishlist_sheet[f'A{row_idx}'] = item.name
            wishlist_sheet[f'B{row_idx}'] = item.description
            wishlist_sheet[f'C{row_idx}'] = float(item.estimated_cost)
            wishlist_sheet[f'D{row_idx}'] = item.priority
            wishlist_sheet[f'E{row_idx}'] = item.created_at.strftime('%Y-%m-%d')
            
            for col in range(1, 6):
                wishlist_sheet.cell(row=row_idx, column=col).border = border
        
        # Adjust column widths
        for sheet in [user_sheet, transactions_sheet, wishlist_sheet]:
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Create the response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="cash_journal_export_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        
        # Save the workbook to the response
        wb.save(response)
        return response

class DeleteAccountView(LoginRequiredMixin, View):
    def post(self, request):
        # Delete all user data
        Transaction.objects.filter(user=request.user).delete()
        WishlistItem.objects.filter(user=request.user).delete()
        UserPreferences.objects.filter(user=request.user).delete()
        
        # Delete the user account
        user = request.user
        logout(request)
        user.delete()
        
        messages.success(request, 'Your account has been successfully deleted.')
        return redirect('login')
